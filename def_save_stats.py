import openpyxl
from openpyxl import Workbook
from class_Person import Person
import os

# create file
#wb = openpyxl.Workbook()

def save_stats(user, time_played, item_acquired):
    # get path of the already existing needed excel file
    path = str(os.path.abspath("Stats_files.xlsx"))
  
    # To open the workbook, workbook object is created 
    wb = openpyxl.load_workbook(path) 

    # activate it
    ws = wb.active
    ws.title = "Player's Statistics"
    ws.sheet_properties.tabColor = "FF5722" #2E4053"

    #Names_columns = ("NAME", "AGE", "TIME PLAYED", "ITEM")
    #ws.append(Names_columns)
    name = user.name
    name_upper = name.upper()
    #       C1: NAME;  C2: AGE;  C3: TIME;    C4: ITEM
    data = (name_upper, int(user.age), time_played, item_acquired)
    ws.append(data)

    row = ws.max_row # total rows 

    # character name
    same_name = 0
    for i in range(1, row + 1): 
        cell_obj = ws.cell(row = i, column = 1)
        if name_upper == cell_obj.value:
            same_name += 1 # add all people who had the same name as user
    
    total_age = 0  
    for i in range(1, row + 1): 
        cell_obj = ws.cell(row = i, column = 2)
        total_age += cell_obj.value # add the ages of all users
    average_age = round(total_age / row, 2) # get the average age of all users

    total_time_played = 0
    for i in range(1, row + 1): 
        cell_obj = ws.cell(row = i, column = 3)
        total_time_played += cell_obj.value   # add time played of all users
    average_time_played = round((total_time_played / row)/60, 3) # get the average time played of users

    same_item = 0
    for i in range(1, row + 1): 
        cell_obj = ws.cell(row = i, column = 4)
        if item_acquired == cell_obj.value:
            same_item += 1 # add all people that had the same item as user

    wb.save(filename="Stats_files.xlsx") # save file 
    return same_name, average_age, average_time_played, same_item

#wb.save(filename="Stats_files.xlsx")
